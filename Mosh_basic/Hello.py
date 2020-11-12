#print("Hello world")
#x=2
#y=9
#print(x+y)
#mylist=[1,2,3,4,4,2,3,6]
#Maxi=max(mylist)
#print(Maxi)
#mylist=[1,2,3,2,4,5,2,9,4]
#maxi=mylist[0]
#for item in mylist:
  #  if item>maxi:
 #       maxi=item
#print(maxi)
#mylist=[1,1,1,1,5]
#for item in mylist:
    #letter=item* '*'
   # print(letter)
#Name=input(": ")
#if len(Name)>5 and len(Name)<15:
 #      print("That is fine")
#else:
 #      print("That is not fine")
#command=""
#started=True
#command=input("->: ").lower()
#if command=="start":
   # if started:
    #    print("The car is already started")
    #else:
     #   print("Car should be started")
#elif command=="stop":
 #   if not started:
  #      print("The car is already stopped")
   # else:
    #    print("car should be stopped")
#elif command=="quit":
 #       print("quit")
#else:
 #print("I do not understand you")
#secret_number=18
#number_of_guess=4
#guess=0
#while guess<number_of_guess:
 #   myguess = int(input("->: "))
  #  guess+=1
   # if myguess==secret_number:
    #    print("Well, You are lucky")
     #   break
#else:
 #   print("Sorry, you are not lucky")
 
#print("☺")
#print("✈")
# write code with function that takea emoji in input and display it in output
###########################################
#def get_emoji(message):
 #   words=message.split(' ')
  #  emoji={
   #           ":)" : "☺"
    #    }
    #output=""
    #for word in words:
     #output+= emoji.get(word,word)+" "
    #return output
#message=input("->:")
#get_emoji(message)
#print(get_emoji(message))

 ##########################################
 # Write a function that takes a number and calculate its square
#def square_calc(number):
 #   result=number*number
  #  return result


#number=int(input(":->"))
#print(square_calc(number))
###########################################
# Write a code which takes the list of integers with repeated values, removes those repeaded values and prints the new list
#list=[1,1,2,1,2,4,5,6]
#final_list=[]
#for item in list:
   # if item not in final_list:
      #  final_list.append(item)
#print(final_list)
##########################################
#Write a code that takes the number as an integer and returns the word format of that number
#my_number=int(input(":-> "))
#if my_number==1:
 #print("one")
#elif  my_number==2:
 #print("two")
#else:
 #print("It does not make sense")
#########################################
# Simple example for how function works
#def greet_user():
 #   print("Hi shahin")

#print("Start")
#greet_user()
#print("Finish")
########################################
#Simple example for how function works 2
#def greet_user(name):
 #   print(f" {name}")


#print("Hi there")
#greet_user("Shahin Ahmadi")
#######################################
# Example for exception. The exception error used to exclude some error from output
#try:
 #   example=int(input("-> "))
  #  print(example)

#except: ValueError
####################################
# Another example for exception
#try:
 #   number=int(input("->"))
  #  income=1000
  #  daily_income=income/number
  #  print(daily_income)

#except:  ZeroDivisionError
###################################
# Getting and including emoji in a sentence(This is already wreitten)
#def get_emoji(message):
 #   words=message.split (' ')
  #  emoji={
       #     ":)":"☺"
    #      }
    #output=""
    #for word in words:
     #output+=emoji.get(word,word)+ " "
    #return output


#message=input("->")
#print(get_emoji(message))
#######################################
# class example; creating type using class
#class  Point:
 #    def draw(self):
  #      print("I love you")
   #  def read(self):
    #   print(" You are the best ")


#point1=Point()
#point1.draw()
#print(point1)

#point2=Point.read()
#print(point2)
#########################################
# Construction
#class  Point:
 #   def __init__(self,x,y):
  #      self.x=x
   #     self.y=y


    #def draw(self):
     #   print("I love you")
     #  print(" You are the best ")


#point1=Point(1,20)
#print(point1.x)
#point1.draw()
#print(point1)
################################################
#class Person:
 #   def __init__(self,name):
  #      self.name=name


   # def talk(self):
    #    print(f"talk {self.name}")



#John=Person("John Smith")
#print(John.name)
#John.talk()
#######################################################
#Inheritence in Python
#class Mammal:
 #   def walk(self):
  #      print("walk")


#class Dog(Mammal):
 # def bark(self):
  #    print("bark")

#class Cat(Mammal):
 #  def be_annoying(self):
  #     print("annoying")

#cat1=Cat()
#cat1.Cat()
#dog1=Dog()s
#dog1.walk()
#########################################################
#Modules
#import convertors
#from convertors import kg_to_lbs
#print(kg_to_lbs(100))

#print(convertors.kg_to_lbs(70))
######################################################
#Excercise for the modules

#from utils import find_max
#numbers=[1,2,3,10,67,2]
#max=find_max(numbers)
#print(max)
########################################################
#Package
#from  ecommerce import  shipping
#shipping.calc_shipping()
########################################################
################Proccessing spreadsheet
# openpyxl as xl
#from openpyxl.chart import BarChart, Reference
#wb=xl.load_workbook('transactions.xlsx')
#sheet=wb['Sheet1']
#cell=sheet['a1']
#cell=sheet.cell(1,1)
#print(cell.value)
#print(sheet.max_row)
#for row in range(2,sheet.max_row+1):
 #   cell=sheet.cell(row,3)
  #  Corrected_price=cell.value*0.9
   # Corrected_price_cell=sheet.cell(row,4)
    #Corrected_price_cell.value=Corrected_price
    #values=Reference(sheet,
     #         min_row=2,
      #        max_row=sheet.max_row,
       #       min_col=4,
        #      max_col=4)
    #chart=BarChart()
    #chart.add_data(values)
    #sheet.add_chart(chart,'e2')
    #wb.save('transactions2.xlsx')

#name=input('What is your name? ')
#print('Hi' + name)
'''command=input("Enter your command: ").lower()
started=False
while command=='help':
    command1=input("Enter your command1: ").lower()
    if command1=='start':
        if started:
            print('The car already is started')
        else:
            started=True
            print("The car should start")
    elif command1=='stop':
        if not started:
            print("The car is stopped already")
        else:
         started=False
         print("The car should stop")

    else:
        print("Quit the program")
else:
    print("Quit the program") '''

#### Code to process spreadsheet ####

''' import openpyxl as xl
from openpyxl.chart import BarChart, Reference
wb=xl.load_workbook('transactions.xlsx')
sheet=wb['Sheet1']
cell=sheet['a1']
cell=sheet.cell(1,1)
print(cell.value)
print(sheet.max_row)
for row in range(2, sheet.max_row+1):
    cell=sheet.cell(row,3)
    corrected_price=cell.value*0.9
    corrected_price_cell=sheet.cell(row,4)
    corrected_price_cell.value=corrected_price
    values= Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)

chart=BarChart()
chart.add_data(values)
sheet.add_chart(chart,'e2')
wb.save('transactions2.xlsx') '''
''' correct_number=9
try_number=4
guess=0
while guess<try_number:
    your_guess = int(input("Your guess:  "))
    guess += 1
    if your_guess==correct_number:
        print("Your guess is true")
        break



else:
         print("Sorry, your guess was not true") '''

''' from utils2 import maximum_calculater
mylist=[1,3,2,5,2,4,5,6,7,234,654,68]
print(maximum_calculater(mylist)) '''

'''your_command=input('Enter your command: ')
started=True
while your_command=='help':
    command1=input('Enter your command: ')
    if command1=='start':
        if started:
            print("The car is already started")
        else:
            print("The car should start")
    elif command1=='stop':
        if not started:
            print('The car is already stopped')
        else:
            print('The car should stop')
else:
    print('Quit the program')'''
'''def square(number):
    result=number*number
    return result
print(square(100))'''
'''list1=[1,2,3,4,4,5,5,5,5,6,6,6,6,123,123,6,5,8,9,0]
print(max(list1))
list2=[]
for i in list1:
    if i not in list2:
     list2.append(i)
print(list2)'''
from convertor import lbs_to_kg
from convertor import kg_to_lbs
weight1=lbs_to_kg(10)
print("This number is in kg: ")
print(weight1)
weight2=kg_to_lbs(10)
print('This number is in lbs: ')
print(weight2)















